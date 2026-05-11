from flask import Flask, render_template, request, jsonify, send_file
from datetime import date, datetime, timedelta
import json, os, io, sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Path resolution: works from source AND from installed/PyInstaller bundle ─
if getattr(sys, 'frozen', False):
    _BUNDLE = sys._MEIPASS
    app = Flask(__name__,
                template_folder=os.path.join(_BUNDLE, 'templates'),
                static_folder=os.path.join(_BUNDLE, 'static'))
    DATA_DIR = os.path.join(os.path.expanduser('~'), '.activity-tracker', 'data')
else:
    app = Flask(__name__)
    _local_data = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')
    # If the app is installed somewhere not writable by this user (e.g. /opt/),
    # store data in the user's home directory instead.
    try:
        os.makedirs(_local_data, exist_ok=True)
        test_f = os.path.join(_local_data, '.write_test')
        open(test_f, 'w').close()
        os.remove(test_f)
        DATA_DIR = _local_data
    except OSError:
        DATA_DIR = os.path.join(os.path.expanduser('~'), '.activity-tracker', 'data')

os.makedirs(DATA_DIR, exist_ok=True)

def data_file(name):
    return os.path.join(DATA_DIR, f"{name}.json")

def load_json(name, default):
    path = data_file(name)
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return default

def save_json(name, data):
    with open(data_file(name), "w") as f:
        json.dump(data, f, indent=2, default=str)

today = lambda: date.today().isoformat()

# ── routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")

# ── Tasks ────────────────────────────────────────────────────────────────────

@app.route("/api/tasks", methods=["GET"])
def get_tasks():
    d = request.args.get("date", today())
    tasks = load_json("tasks", {})
    return jsonify(tasks.get(d, []))

@app.route("/api/tasks", methods=["POST"])
def add_task():
    d = request.json.get("date", today())
    tasks = load_json("tasks", {})
    day_tasks = tasks.get(d, [])
    task = {
        "id": int(datetime.now().timestamp() * 1000),
        "text": request.json["text"],
        "priority": request.json.get("priority", "Medium"),
        "category": request.json.get("category", "General"),
        "done": False,
        "created": datetime.now().isoformat()
    }
    day_tasks.append(task)
    tasks[d] = day_tasks
    save_json("tasks", tasks)
    return jsonify(task)

@app.route("/api/tasks/<int:tid>", methods=["PATCH"])
def update_task(tid):
    d = request.json.get("date", today())
    tasks = load_json("tasks", {})
    for t in tasks.get(d, []):
        if t["id"] == tid:
            t.update({k: v for k, v in request.json.items() if k != "date"})
    save_json("tasks", tasks)
    return jsonify({"ok": True})

@app.route("/api/tasks/<int:tid>", methods=["DELETE"])
def delete_task(tid):
    d = request.args.get("date", today())
    tasks = load_json("tasks", {})
    tasks[d] = [t for t in tasks.get(d, []) if t["id"] != tid]
    save_json("tasks", tasks)
    return jsonify({"ok": True})

# ── Habits ───────────────────────────────────────────────────────────────────

@app.route("/api/habits/list", methods=["GET"])
def get_habit_list():
    return jsonify(load_json("habit_list", [
        {"id": 1, "name": "Exercise", "target": 1, "unit": "session", "icon": "🏃"},
        {"id": 2, "name": "Water",    "target": 8, "unit": "glasses", "icon": "💧"},
        {"id": 3, "name": "Reading",  "target": 30,"unit": "minutes", "icon": "📚"},
        {"id": 4, "name": "Sleep",    "target": 8, "unit": "hours",   "icon": "😴"},
        {"id": 5, "name": "Meditation","target": 10,"unit": "minutes","icon": "🧘"},
    ]))

@app.route("/api/habits/list", methods=["POST"])
def add_habit():
    habits = load_json("habit_list", [])
    h = request.json
    h["id"] = int(datetime.now().timestamp() * 1000)
    habits.append(h)
    save_json("habit_list", habits)
    return jsonify(h)

@app.route("/api/habits/list/<int:hid>", methods=["DELETE"])
def delete_habit(hid):
    habits = load_json("habit_list", [])
    save_json("habit_list", [h for h in habits if h["id"] != hid])
    return jsonify({"ok": True})

@app.route("/api/habits/log", methods=["GET"])
def get_habit_log():
    d = request.args.get("date", today())
    logs = load_json("habit_log", {})
    return jsonify(logs.get(d, {}))

@app.route("/api/habits/log", methods=["POST"])
def set_habit_log():
    d = request.json.get("date", today())
    hid = str(request.json["habit_id"])
    val = request.json["value"]
    logs = load_json("habit_log", {})
    if d not in logs:
        logs[d] = {}
    logs[d][hid] = val
    save_json("habit_log", logs)
    return jsonify({"ok": True})

# ── Work Hours ────────────────────────────────────────────────────────────────

@app.route("/api/work", methods=["GET"])
def get_work():
    d = request.args.get("date", today())
    work = load_json("work", {})
    return jsonify(work.get(d, []))

@app.route("/api/work", methods=["POST"])
def add_work():
    d = request.json.get("date", today())
    work = load_json("work", {})
    entry = {
        "id": int(datetime.now().timestamp() * 1000),
        "project": request.json["project"],
        "hours": float(request.json["hours"]),
        "notes": request.json.get("notes", ""),
        "created": datetime.now().isoformat()
    }
    work.setdefault(d, []).append(entry)
    save_json("work", work)
    return jsonify(entry)

@app.route("/api/work/<int:wid>", methods=["PATCH"])
def update_work(wid):
    d = request.json.get("date", today())
    work = load_json("work", {})
    for w in work.get(d, []):
        if w["id"] == wid:
            for k, v in request.json.items():
                if k != "date":
                    w[k] = v
    save_json("work", work)
    return jsonify({"ok": True})

@app.route("/api/work/<int:wid>", methods=["DELETE"])
def delete_work(wid):
    d = request.args.get("date", today())
    work = load_json("work", {})
    work[d] = [w for w in work.get(d, []) if w["id"] != wid]
    save_json("work", work)
    return jsonify({"ok": True})

# ── Goals ─────────────────────────────────────────────────────────────────────

@app.route("/api/goals", methods=["GET"])
def get_goals():
    return jsonify(load_json("goals", []))

@app.route("/api/goals", methods=["POST"])
def add_goal():
    goals = load_json("goals", [])
    g = {
        "id": int(datetime.now().timestamp() * 1000),
        "title": request.json["title"],
        "target": float(request.json["target"]),
        "current": float(request.json.get("current", 0)),
        "unit": request.json.get("unit", ""),
        "deadline": request.json.get("deadline", ""),
        "category": request.json.get("category", "Personal"),
        "created": today()
    }
    goals.append(g)
    save_json("goals", goals)
    return jsonify(g)

@app.route("/api/goals/<int:gid>", methods=["PATCH"])
def update_goal(gid):
    goals = load_json("goals", [])
    for g in goals:
        if g["id"] == gid:
            g.update({k: v for k, v in request.json.items()})
    save_json("goals", goals)
    return jsonify({"ok": True})

@app.route("/api/goals/<int:gid>", methods=["DELETE"])
def delete_goal(gid):
    goals = load_json("goals", [])
    save_json("goals", [g for g in goals if g["id"] != gid])
    return jsonify({"ok": True})

# ── Diet ──────────────────────────────────────────────────────────────────────

@app.route("/api/diet", methods=["GET"])
def get_diet():
    d = request.args.get("date", today())
    diet = load_json("diet", {})
    return jsonify(diet.get(d, []))

@app.route("/api/diet", methods=["POST"])
def add_diet():
    d = request.json.get("date", today())
    diet = load_json("diet", {})
    entry = {
        "id": int(datetime.now().timestamp() * 1000),
        "meal":     request.json["meal"],
        "food":     request.json["food"],
        "calories": int(request.json.get("calories", 0)),
        "protein":  float(request.json.get("protein", 0)),
        "carbs":    float(request.json.get("carbs", 0)),
        "fat":      float(request.json.get("fat", 0)),
        "fiber":    float(request.json.get("fiber", 0)),
        "sugar":    float(request.json.get("sugar", 0)),
        "sodium":   float(request.json.get("sodium", 0)),
        "time":     request.json.get("time", "")
    }
    diet.setdefault(d, []).append(entry)
    save_json("diet", diet)
    return jsonify(entry)

@app.route("/api/diet/<int:did>", methods=["PATCH"])
def update_diet(did):
    d = request.json.get("date", today())
    diet = load_json("diet", {})
    for x in diet.get(d, []):
        if x["id"] == did:
            for k, v in request.json.items():
                if k != "date":
                    x[k] = v
    save_json("diet", diet)
    return jsonify({"ok": True})

@app.route("/api/diet/<int:did>", methods=["DELETE"])
def delete_diet(did):
    d = request.args.get("date", today())
    diet = load_json("diet", {})
    diet[d] = [x for x in diet.get(d, []) if x["id"] != did]
    save_json("diet", diet)
    return jsonify({"ok": True})

@app.route("/api/diet/plan", methods=["GET"])
def get_diet_plan():
    return jsonify(load_json("diet_plan", {
        "calories": 2000, "protein": 150, "carbs": 250,
        "fat": 65, "fiber": 30, "sugar": 50, "sodium": 2300
    }))

@app.route("/api/diet/plan", methods=["POST"])
def save_diet_plan():
    plan = {
        "calories": int(request.json.get("calories", 2000)),
        "protein":  float(request.json.get("protein", 150)),
        "carbs":    float(request.json.get("carbs", 250)),
        "fat":      float(request.json.get("fat", 65)),
        "fiber":    float(request.json.get("fiber", 30)),
        "sugar":    float(request.json.get("sugar", 50)),
        "sodium":   float(request.json.get("sodium", 2300)),
    }
    save_json("diet_plan", plan)
    return jsonify(plan)

# ── Body Profile ─────────────────────────────────────────────────────────────

@app.route("/api/profile", methods=["GET"])
def get_profile():
    return jsonify(load_json("profile", {
        "name": "", "age": 25, "gender": "male",
        "height": 175, "weight": 70, "target_weight": 70,
        "activity": "moderate", "goal": "maintain", "unit_system": "metric"
    }))

@app.route("/api/profile", methods=["POST"])
def save_profile():
    p = {
        "name":          request.json.get("name", ""),
        "age":           int(request.json.get("age", 25)),
        "gender":        request.json.get("gender", "male"),
        "height":        float(request.json.get("height", 175)),
        "weight":        float(request.json.get("weight", 70)),
        "target_weight": float(request.json.get("target_weight", request.json.get("weight", 70))),
        "activity":      request.json.get("activity", "moderate"),
        "goal":          request.json.get("goal", "maintain"),
        "unit_system":   request.json.get("unit_system", "metric"),
    }
    save_json("profile", p)
    tdee_val  = _tdee(p)
    macros    = _suggest_macros(p)
    bmr       = _bmr(p)
    return jsonify({**p, "tdee": tdee_val, "bmr": bmr, "macros": macros,
                    "maintenance": round(bmr * {"sedentary":1.2,"light":1.375,"moderate":1.55,"active":1.725,"extra":1.9}.get(p["activity"],1.55))})

def _bmr(p):
    w, h, a, g = p["weight"], p["height"], p["age"], p["gender"]
    return round(10*w + 6.25*h - 5*a + (5 if g == "male" else -161))

def _tdee(p):
    mult = {"sedentary":1.2,"light":1.375,"moderate":1.55,"active":1.725,"extra":1.9}
    return round(_bmr(p) * mult.get(p["activity"], 1.55))

def _suggest_macros(p):
    tdee = _tdee(p)
    goal, w = p["goal"], p["weight"]
    adj  = {"lose_weight":-500,"maintain":0,"gain_muscle":300,"athletic":200}
    cals = tdee + adj.get(goal, 0)
    pro  = {"lose_weight":2.2,"maintain":1.8,"gain_muscle":2.5,"athletic":2.0}
    fat  = {"lose_weight":0.9,"maintain":1.0,"gain_muscle":1.0,"athletic":1.2}
    p_g  = round(pro.get(goal, 1.8) * w)
    f_g  = round(fat.get(goal, 1.0) * w)
    c_g  = round(max(0, (cals - p_g*4 - f_g*9) / 4))
    return {"calories": cals, "protein": p_g, "carbs": c_g, "fat": f_g,
            "fiber": 30, "sugar": 45, "sodium": 2300}

# ── Weight Log ────────────────────────────────────────────────────────────────

@app.route("/api/weight", methods=["GET"])
def get_weight():
    days = int(request.args.get("days", 30))
    log  = load_json("weight_log", {})
    end  = date.today()
    out  = []
    for i in range(days-1, -1, -1):
        d = (end - timedelta(days=i)).isoformat()
        if d in log:
            out.append({"date": d, "weight": log[d]["weight"],
                        "notes": log[d].get("notes","")})
    return jsonify(out)

@app.route("/api/weight", methods=["POST"])
def log_weight():
    d = request.json.get("date", today())
    log = load_json("weight_log", {})
    entry = {"weight": float(request.json["weight"]),
             "notes":  request.json.get("notes",""),
             "logged": datetime.now().isoformat()}
    log[d] = entry
    save_json("weight_log", log)
    return jsonify({**entry, "date": d})

@app.route("/api/weight/<string:d>", methods=["DELETE"])
def delete_weight(d):
    log = load_json("weight_log", {})
    log.pop(d, None)
    save_json("weight_log", log)
    return jsonify({"ok": True})

# ── Diet Plans Library ────────────────────────────────────────────────────────

@app.route("/api/diet/plans", methods=["GET"])
def get_diet_plans():
    return jsonify(load_json("diet_plans", []))

@app.route("/api/diet/plans", methods=["POST"])
def create_diet_plan():
    plans = load_json("diet_plans", [])
    plan  = {
        "id":       int(datetime.now().timestamp()*1000),
        "name":     request.json.get("name","My Plan"),
        "goal":     request.json.get("goal","maintain"),
        "month":    request.json.get("month", date.today().strftime("%Y-%m")),
        "calories": int(request.json.get("calories",2000)),
        "protein":  float(request.json.get("protein",150)),
        "carbs":    float(request.json.get("carbs",250)),
        "fat":      float(request.json.get("fat",65)),
        "fiber":    float(request.json.get("fiber",30)),
        "sugar":    float(request.json.get("sugar",50)),
        "sodium":   float(request.json.get("sodium",2300)),
        "notes":    request.json.get("notes",""),
        "active":   False,
        "created":  date.today().isoformat(),
    }
    if not plans:
        plan["active"] = True
    plans.append(plan)
    save_json("diet_plans", plans)
    _sync_active_plan(plans)
    return jsonify(plan)

@app.route("/api/diet/plans/<int:pid>", methods=["DELETE"])
def delete_diet_plan(pid):
    plans = [p for p in load_json("diet_plans",[]) if p["id"] != pid]
    save_json("diet_plans", plans)
    return jsonify({"ok": True})

@app.route("/api/diet/plans/<int:pid>/activate", methods=["POST"])
def activate_diet_plan(pid):
    plans = load_json("diet_plans", [])
    for p in plans:
        p["active"] = (p["id"] == pid)
    save_json("diet_plans", plans)
    _sync_active_plan(plans)
    return jsonify({"ok": True})

def _sync_active_plan(plans):
    active = next((p for p in plans if p.get("active")), None)
    if active:
        save_json("diet_plan", {k: active[k] for k in
            ["calories","protein","carbs","fat","fiber","sugar","sodium"]})

# ── Adherence Engine ──────────────────────────────────────────────────────────

@app.route("/api/adherence")
def get_adherence():
    days   = int(request.args.get("days", 30))
    plan   = load_json("diet_plan", {"calories":2000,"protein":150,"carbs":250,
                                     "fat":65,"fiber":30,"sugar":50,"sodium":2300})
    diet   = load_json("diet",  {})
    tasks  = load_json("tasks", {})
    habits = load_json("habit_log", {})
    h_list = load_json("habit_list", [])
    end, out = date.today(), []
    for i in range(days-1, -1, -1):
        d  = (end - timedelta(days=i)).isoformat()
        de = diet.get(d, [])
        te = tasks.get(d, [])
        he = habits.get(d, {})
        if not de and not te:
            out.append({"date": d, "logged": False}); continue
        totals = {k: sum(e.get(k,0) for e in de)
                  for k in ["calories","protein","carbs","fat","fiber","sugar","sodium"]}
        task_pct = (sum(1 for t in te if t.get("done")) / len(te) * 100) if te else 0
        hab_pct  = 0
        if h_list:
            scores = []
            for h in h_list:
                val = he.get(str(h["id"]), 0)
                scores.append(min(val / (h.get("target") or 1), 1.0))
            hab_pct = sum(scores)/len(scores)*100
        out.append({
            "date":    d, "logged": True,
            "calories_pct":  round(totals["calories"] / plan["calories"] * 100) if plan["calories"] else 0,
            "protein_pct":   round(totals["protein"]  / plan["protein"]  * 100) if plan["protein"]  else 0,
            "carbs_pct":     round(totals["carbs"]    / plan["carbs"]    * 100) if plan["carbs"]    else 0,
            "fat_pct":       round(totals["fat"]      / plan["fat"]      * 100) if plan["fat"]      else 0,
            "fiber_pct":     round(totals["fiber"]    / plan["fiber"]    * 100) if plan["fiber"]    else 0,
            "sugar_pct":     round(totals["sugar"]    / plan["sugar"]    * 100) if plan["sugar"]    else 0,
            "sodium_pct":    round(totals["sodium"]   / plan["sodium"]   * 100) if plan["sodium"]   else 0,
            "task_pct":      round(task_pct),
            "habit_pct":     round(hab_pct),
            "calories":      totals["calories"],
        })
    return jsonify(out)

# ── Coaching Engine ───────────────────────────────────────────────────────────

@app.route("/api/coaching")
def get_coaching():
    adh_days = 14
    plan     = load_json("diet_plan", {})
    profile  = load_json("profile", {"goal":"maintain","weight":70})
    diet     = load_json("diet", {})
    end = date.today()
    adh = []
    for i in range(adh_days-1, -1, -1):
        d  = (end - timedelta(days=i)).isoformat()
        de = diet.get(d, [])
        if not de:
            adh.append({"logged": False}); continue
        totals = {k: sum(e.get(k,0) for e in de)
                  for k in ["calories","protein","fiber","sugar","sodium"]}
        adh.append({
            "logged": True,
            "cal_pct":    totals["calories"] / (plan.get("calories",2000) or 2000) * 100,
            "pro_pct":    totals["protein"]  / (plan.get("protein",150)   or 150)  * 100,
            "fiber_pct":  totals["fiber"]    / (plan.get("fiber",30)      or 30)   * 100,
            "sugar_pct":  totals["sugar"]    / (plan.get("sugar",50)      or 50)   * 100,
            "sodium_pct": totals["sodium"]   / (plan.get("sodium",2300)   or 2300) * 100,
        })
    logged   = [d for d in adh if d["logged"]]
    recent7  = [d for d in adh[-7:] if d["logged"]]
    goal     = profile.get("goal","maintain")
    insights = []

    def avg(lst, key): return sum(d[key] for d in lst) / len(lst) if lst else 0

    # Tracking consistency
    days_logged = len(recent7)
    if days_logged < 4:
        insights.append({"type":"warning","icon":"📅","priority":10,
            "title":"Track Every Meal",
            "message":f"You only logged meals on {days_logged}/7 days this week. Consistent tracking is the #1 predictor of reaching your goal.",
            "action":"Set a phone reminder to log meals at breakfast, lunch and dinner."})
    elif days_logged == 7:
        insights.append({"type":"success","icon":"🔥","priority":2,
            "title":"Perfect Tracking Streak!",
            "message":"7/7 days logged this week — you're building a champion's habit.",
            "action":None})

    # Calorie analysis
    if recent7:
        cal = avg(recent7, "cal_pct")
        if goal == "lose_weight":
            if cal > 110:
                insights.append({"type":"alert","icon":"🔴","priority":9,
                    "title":"Caloric Surplus — Fat Loss Stalled",
                    "message":f"You're averaging {cal:.0f}% of your calorie target. Even a consistent 200 kcal surplus adds up to fat gain.",
                    "action":"Reduce dinner by one palm-size portion or skip the evening snack."})
            elif cal < 75:
                insights.append({"type":"warning","icon":"⚠️","priority":8,
                    "title":"Deficit Too Deep",
                    "message":"Eating under 75% of target triggers muscle breakdown and metabolic adaptation.",
                    "action":"Add a high-protein snack: Greek yogurt, cottage cheese, or a boiled egg."})
            elif 82 <= cal <= 100:
                insights.append({"type":"success","icon":"✅","priority":2,
                    "title":"Ideal Fat-Loss Deficit",
                    "message":f"Averaging {cal:.0f}% — perfect for losing 0.5–1 kg/week without muscle loss.",
                    "action":None})
        elif goal == "gain_muscle":
            if cal < 95:
                insights.append({"type":"warning","icon":"📈","priority":8,
                    "title":"Not Eating Enough to Grow",
                    "message":f"Averaging {cal:.0f}% of your surplus target. Muscles need fuel to rebuild bigger.",
                    "action":"Add a pre-bed snack: casein shake or cottage cheese with fruit."})
            elif cal > 115:
                insights.append({"type":"warning","icon":"⚠️","priority":6,
                    "title":"Surplus Too Large",
                    "message":"A surplus over 15% leads to excessive fat gain. Aim for 200–300 kcal above TDEE.",
                    "action":"Trim one high-carb portion at lunch or dinner."})
        elif goal == "maintain":
            if cal > 110 or cal < 85:
                insights.append({"type":"warning","icon":"⚖️","priority":7,
                    "title":"Caloric Balance Off",
                    "message":f"Maintenance requires hitting close to 100%. You're averaging {cal:.0f}%.",
                    "action":"Use a food scale for 3 days to recalibrate your portion sense."})

    # Protein
    if recent7:
        pro = avg(recent7, "pro_pct")
        if pro < 75:
            insights.append({"type":"alert","icon":"💪","priority":8,
                "title":"Protein Too Low",
                "message":f"Only {pro:.0f}% of your protein target. Low protein causes muscle loss, hunger, and slower metabolism.",
                "action":"Add a protein source to every meal: chicken, fish, eggs, legumes, or whey."})
        elif pro >= 100:
            insights.append({"type":"success","icon":"🏆","priority":1,
                "title":"Protein Goal Nailed!",
                "message":f"Averaging {pro:.0f}% of target — your muscles have everything they need to recover and grow.",
                "action":None})

    # Sugar
    if recent7:
        sug = avg(recent7, "sugar_pct")
        if sug > 130:
            insights.append({"type":"alert","icon":"🍬","priority":7,
                "title":"Sugar Intake Way Over Limit",
                "message":f"Averaging {sug:.0f}% of your sugar limit. Excess sugar drives fat storage, inflammation, and energy crashes.",
                "action":"Swap sugary snacks for berries, dark chocolate, or unsweetened Greek yogurt."})

    # Fiber
    if recent7:
        fib = avg(recent7, "fiber_pct")
        if fib < 60:
            insights.append({"type":"info","icon":"🌿","priority":5,
                "title":"Boost Your Fiber",
                "message":f"Only {fib:.0f}% of fiber target. Fiber keeps you full, regulates blood sugar, and feeds gut bacteria.",
                "action":"Add oats, lentils, broccoli, chia seeds, or an apple with skin daily."})

    # Sodium
    if recent7:
        sod = avg(recent7, "sodium_pct")
        if sod > 120:
            insights.append({"type":"info","icon":"🧂","priority":4,
                "title":"Watch Your Sodium",
                "message":f"Averaging {sod:.0f}% of sodium limit. High sodium causes water retention and elevated blood pressure.",
                "action":"Cook from scratch more often and reduce sauces, pickles, and processed foods."})

    # Overall weekly score
    if recent7:
        cal_s = min(avg(recent7,"cal_pct")/100,1.2)
        pro_s = min(avg(recent7,"pro_pct")/100,1.0)
        fib_s = min(avg(recent7,"fiber_pct")/100,1.0)
        sug_s = max(0, 1 - max(avg(recent7,"sugar_pct")-100,0)/100)
        score = round((cal_s*0.3 + pro_s*0.3 + fib_s*0.2 + sug_s*0.1 + (days_logged/7)*0.1)*100)
    else:
        score = 0

    insights.sort(key=lambda x: x["priority"], reverse=True)
    return jsonify({"insights": insights[:6], "score": score,
                    "days_logged": days_logged, "streak": _streak()})

def _streak():
    diet = load_json("diet", {})
    s, d = 0, date.today()
    while True:
        if d.isoformat() not in diet: break
        s += 1; d -= timedelta(days=1)
    return s

# ── Plan Suggestion (Month Review) ────────────────────────────────────────────

@app.route("/api/coaching/suggest-plan", methods=["GET"])
def suggest_plan():
    profile = load_json("profile", {"goal":"maintain","weight":70})
    plan    = load_json("diet_plan", {})
    diet    = load_json("diet", {})
    wlog    = load_json("weight_log", {})
    end     = date.today()
    days_data = []
    for i in range(29, -1, -1):
        d  = (end - timedelta(days=i)).isoformat()
        de = diet.get(d, [])
        if de:
            days_data.append({
                "calories": sum(e.get("calories",0) for e in de),
                "protein":  sum(e.get("protein",0)  for e in de),
            })
    if not days_data:
        return jsonify({"suggestion":None, "reason":"Not enough data yet. Log meals for at least 7 days."})
    logged_n  = len(days_data)
    avg_cal   = sum(d["calories"] for d in days_data) / logged_n
    avg_pro   = sum(d["protein"]  for d in days_data) / logged_n
    tgt_cal   = plan.get("calories", 2000)
    tgt_pro   = plan.get("protein",  150)
    cal_adh   = avg_cal / tgt_cal * 100 if tgt_cal else 0
    pro_adh   = avg_pro / tgt_pro  * 100 if tgt_pro  else 0
    goal      = profile.get("goal","maintain")
    w_list    = sorted(wlog.items())
    w_change  = None
    if len(w_list) >= 2:
        w_change = round(w_list[-1][1]["weight"] - w_list[0][1]["weight"], 1)
    suggested = dict(plan)
    reasons   = []
    if goal == "lose_weight":
        if w_change is not None and w_change > -0.5:
            suggested["calories"] = max(1200, tgt_cal - 150)
            reasons.append(f"Weight dropped only {w_change:+.1f} kg — reducing calories by 150 kcal.")
        elif cal_adh > 105:
            suggested["calories"] = max(1200, tgt_cal - 100)
            reasons.append(f"Averaging {cal_adh:.0f}% of calorie target — tightening by 100 kcal.")
        if pro_adh < 80:
            suggested["protein"] = round(tgt_pro * 1.1)
            reasons.append(f"Protein adherence was {pro_adh:.0f}% — raising target by 10%.")
        if not reasons:
            reasons.append(f"Great adherence ({cal_adh:.0f}% cal, {pro_adh:.0f}% protein). Maintaining current plan.")
    elif goal == "gain_muscle":
        if w_change is not None and w_change < 0.5:
            suggested["calories"] = tgt_cal + 150
            reasons.append(f"Only {w_change:+.1f} kg gained — adding 150 kcal to accelerate growth.")
        if pro_adh >= 95:
            suggested["protein"] = round(tgt_pro * 1.05)
            reasons.append("Consistently hitting protein — nudging up by 5% to support more muscle.")
        if not reasons:
            reasons.append(f"Good bulk progress ({cal_adh:.0f}% cal). Maintaining surplus.")
    else:
        if abs(cal_adh - 100) < 5 and pro_adh >= 85:
            reasons.append("Excellent adherence — no changes needed. Keep it up!")
        elif cal_adh > 105:
            suggested["calories"] = tgt_cal - 100
            reasons.append(f"Slight surplus at {cal_adh:.0f}% — trimming 100 kcal.")
        elif cal_adh < 85:
            suggested["calories"] = tgt_cal + 100
            reasons.append(f"Under-eating at {cal_adh:.0f}% — adding 100 kcal.")
    next_month = (date.today().replace(day=1) + timedelta(days=32)).strftime("%Y-%m")
    return jsonify({
        "suggestion": suggested,
        "reasons":    reasons,
        "stats": {"days_logged": logged_n, "avg_cal": round(avg_cal),
                  "avg_protein": round(avg_pro), "cal_adherence": round(cal_adh),
                  "pro_adherence": round(pro_adh), "weight_change": w_change},
        "next_month": next_month,
    })

# ── Week Summary ─────────────────────────────────────────────────────────────

@app.route("/api/summary/week")
def week_summary():
    all_tasks = load_json("tasks", {})
    all_work  = load_json("work", {})
    all_diet  = load_json("diet", {})
    result = []
    for i in range(6, -1, -1):
        day = (date.today() - timedelta(days=i)).isoformat()
        dt = all_tasks.get(day, [])
        dw = all_work.get(day, [])
        dd = all_diet.get(day, [])
        result.append({
            "date":    day,
            "label":   day[5:],
            "done":    sum(1 for t in dt if t.get("done")),
            "total":   len(dt),
            "hours":   round(sum(w["hours"] for w in dw), 2),
            "calories": sum(e.get("calories", 0) for e in dd),
        })
    return jsonify(result)

# ── Notes / Journal ──────────────────────────────────────────────────────────

@app.route("/api/notes", methods=["GET"])
def get_notes():
    d = request.args.get("date", today())
    return jsonify(load_json("notes", {}).get(d, []))

@app.route("/api/notes", methods=["POST"])
def add_note():
    d = request.json.get("date", today())
    notes = load_json("notes", {})
    note = {
        "id":      int(datetime.now().timestamp() * 1000),
        "type":    request.json.get("type", "note"),
        "title":   request.json.get("title", ""),
        "content": request.json.get("content", ""),
        "mood":    request.json.get("mood", ""),
        "tags":    request.json.get("tags", []),
        "pinned":  False,
        "created": datetime.now().isoformat(),
        "updated": datetime.now().isoformat(),
    }
    notes.setdefault(d, []).insert(0, note)
    save_json("notes", notes)
    return jsonify(note)

@app.route("/api/notes/<int:nid>", methods=["PATCH"])
def update_note(nid):
    d = request.json.get("date", today())
    notes = load_json("notes", {})
    for n in notes.get(d, []):
        if n["id"] == nid:
            n.update({k: v for k, v in request.json.items() if k != "date"})
            n["updated"] = datetime.now().isoformat()
    save_json("notes", notes)
    return jsonify({"ok": True})

@app.route("/api/notes/<int:nid>", methods=["DELETE"])
def delete_note(nid):
    d = request.args.get("date", today())
    notes = load_json("notes", {})
    notes[d] = [n for n in notes.get(d, []) if n["id"] != nid]
    save_json("notes", notes)
    return jsonify({"ok": True})

@app.route("/api/notes/search")
def search_notes():
    q = request.args.get("q", "").lower().strip()
    if not q:
        return jsonify([])
    all_notes = load_json("notes", {})
    results = []
    for day, day_notes in sorted(all_notes.items(), reverse=True):
        for n in day_notes:
            if (q in n.get("title","").lower() or
                q in n.get("content","").lower() or
                any(q in t.lower() for t in n.get("tags", []))):
                results.append({**n, "date": day})
    return jsonify(results[:50])

# ── Weekly Summary ───────────────────────────────────────────────────────────

@app.route("/api/summary/weekly")
def weekly_summary():
    import calendar as cal_mod
    d        = request.args.get("date", today())
    dt       = date.fromisoformat(d)
    wstart   = dt - timedelta(days=dt.weekday())   # Monday

    all_tasks   = load_json("tasks", {})
    all_work    = load_json("work", {})
    all_habits  = load_json("habit_log", {})
    all_diet    = load_json("diet", {})
    habit_list  = load_json("habit_list", [])

    days = []
    tot_done = tot_tasks = tot_hours = tot_cal = 0
    work_by_project = {}

    for i in range(7):
        day    = (wstart + timedelta(days=i)).isoformat()
        dtasks = all_tasks.get(day, [])
        dwork  = all_work.get(day, [])
        ddiet  = all_diet.get(day, [])
        dhab   = all_habits.get(day, {})

        done  = sum(1 for t in dtasks if t.get("done"))
        total = len(dtasks)
        hours = round(sum(w["hours"] for w in dwork), 2)
        cal   = sum(e.get("calories", 0) for e in ddiet)

        pcts = [min(float(dhab.get(str(h["id"]), 0)) / float(h.get("target", 1)) * 100, 100)
                for h in habit_list if float(h.get("target", 1)) > 0]
        hab_score = round(sum(pcts) / len(pcts)) if pcts else 0

        for w in dwork:
            work_by_project[w["project"]] = round(
                work_by_project.get(w["project"], 0) + w["hours"], 2)

        tot_done  += done
        tot_tasks += total
        tot_hours += hours
        tot_cal   += cal

        days.append({
            "date": day, "label": (wstart + timedelta(days=i)).strftime("%a"),
            "short": (wstart + timedelta(days=i)).strftime("%m/%d"),
            "done": done, "total": total, "hours": hours,
            "calories": cal, "habit_score": hab_score,
        })

    habit_breakdown = []
    for h in habit_list:
        pcts, completed = [], 0
        for i in range(7):
            day  = (wstart + timedelta(days=i)).isoformat()
            dhab = all_habits.get(day, {})
            logged = float(dhab.get(str(h["id"]), 0))
            tgt    = float(h.get("target", 1))
            pct    = min(logged / tgt * 100, 100) if tgt else 0
            pcts.append(pct)
            if pct >= 100: completed += 1
        habit_breakdown.append({
            "name": h["name"], "icon": h.get("icon", ""),
            "days_completed": completed,
            "avg_pct": round(sum(pcts) / len(pcts)) if pcts else 0,
        })

    all_scores = [d["habit_score"] for d in days]
    return jsonify({
        "week_start": wstart.isoformat(),
        "week_end":   (wstart + timedelta(days=6)).isoformat(),
        "days": days,
        "totals": {
            "done": tot_done, "total": tot_tasks,
            "hours": round(tot_hours, 2), "calories": tot_cal,
            "avg_habit_score": round(sum(all_scores) / len(all_scores)) if all_scores else 0,
        },
        "work_by_project": work_by_project,
        "habit_breakdown": habit_breakdown,
    })

# ── Monthly Summary ───────────────────────────────────────────────────────────

@app.route("/api/summary/monthly")
def monthly_summary():
    import calendar as cal_mod
    d     = request.args.get("date", today())
    dt    = date.fromisoformat(d)
    year, month = dt.year, dt.month
    _, dim      = cal_mod.monthrange(year, month)
    mstart      = date(year, month, 1)

    all_tasks   = load_json("tasks", {})
    all_work    = load_json("work", {})
    all_habits  = load_json("habit_log", {})
    all_diet    = load_json("diet", {})
    habit_list  = load_json("habit_list", [])

    days = []
    work_by_project = {}
    tot_done = tot_tasks = tot_hours = tot_cal = 0

    for i in range(dim):
        day    = (mstart + timedelta(days=i)).isoformat()
        dtasks = all_tasks.get(day, [])
        dwork  = all_work.get(day, [])
        ddiet  = all_diet.get(day, [])
        dhab   = all_habits.get(day, {})

        done  = sum(1 for t in dtasks if t.get("done"))
        total = len(dtasks)
        hours = round(sum(w["hours"] for w in dwork), 2)
        cal   = sum(e.get("calories", 0) for e in ddiet)

        pcts = [min(float(dhab.get(str(h["id"]), 0)) / float(h.get("target", 1)) * 100, 100)
                for h in habit_list if float(h.get("target", 1)) > 0]
        hab_score = round(sum(pcts) / len(pcts)) if pcts else 0

        for w in dwork:
            work_by_project[w["project"]] = round(
                work_by_project.get(w["project"], 0) + w["hours"], 2)

        tot_done  += done
        tot_tasks += total
        tot_hours += hours
        tot_cal   += cal

        days.append({
            "date": day, "day_num": i + 1,
            "weekday": (mstart + timedelta(days=i)).weekday(),
            "done": done, "total": total, "hours": hours,
            "calories": cal, "habit_score": hab_score,
            "completion_pct": round(done / total * 100) if total else 0,
        })

    # Weekly buckets (for the bar chart)
    weeks = []
    for w in range(0, dim, 7):
        chunk = days[w:w + 7]
        weeks.append({
            "label": f"W{w//7+1}",
            "done":  sum(d["done"] for d in chunk),
            "total": sum(d["total"] for d in chunk),
            "hours": round(sum(d["hours"] for d in chunk), 2),
        })

    habit_breakdown = []
    for h in habit_list:
        pcts, completed = [], 0
        for i in range(dim):
            day  = (mstart + timedelta(days=i)).isoformat()
            dhab = all_habits.get(day, {})
            logged = float(dhab.get(str(h["id"]), 0))
            tgt    = float(h.get("target", 1))
            pct    = min(logged / tgt * 100, 100) if tgt else 0
            pcts.append(pct)
            if pct >= 100: completed += 1
        habit_breakdown.append({
            "name": h["name"], "icon": h.get("icon", ""),
            "days_completed": completed, "days_in_month": dim,
            "avg_pct": round(sum(pcts) / len(pcts)) if pcts else 0,
        })

    all_scores = [d["habit_score"] for d in days]
    return jsonify({
        "year": year, "month": month,
        "month_name": mstart.strftime("%B %Y"),
        "days_in_month": dim,
        "month_start_weekday": mstart.weekday(),
        "days": days, "weeks": weeks,
        "totals": {
            "done": tot_done, "total": tot_tasks,
            "hours": round(tot_hours, 2), "calories": tot_cal,
            "avg_habit_score": round(sum(all_scores) / len(all_scores)) if all_scores else 0,
        },
        "work_by_project": work_by_project,
        "habit_breakdown": habit_breakdown,
    })

# ── Excel Export ──────────────────────────────────────────────────────────────

def style_header(ws, row, col, text, bg="1E3A5F", fg="FFFFFF", bold=True, size=12):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def pct_bar(ws, row, col, pct):
    """Write a text-based progress bar."""
    filled = int(pct / 10)
    bar = "█" * filled + "░" * (10 - filled)
    cell = ws.cell(row=row, column=col, value=f"{bar} {pct:.0f}%")
    cell.font = Font(color="2E86AB" if pct < 100 else "27AE60", bold=True)
    cell.alignment = Alignment(horizontal="center")

def add_thick_border(ws, min_row, min_col, max_row, max_col):
    thick = Side(style="medium", color="1E3A5F")
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            b = Border(
                left=thick if cell.column == min_col else Side(style="thin", color="CCCCCC"),
                right=thick if cell.column == max_col else Side(style="thin", color="CCCCCC"),
                top=thick if cell.row == min_row else Side(style="thin", color="CCCCCC"),
                bottom=thick if cell.row == max_row else Side(style="thin", color="CCCCCC"),
            )
            cell.border = b

# ── Body Measurements ────────────────────────────────────────────────────────

MEASUREMENT_FIELDS = ["weight","body_fat","waist","chest","hips","neck","bicep","thigh","calf"]

@app.route("/api/measurements", methods=["GET"])
def get_measurements():
    log = load_json("measurements", {})
    entries = sorted(
        [{"date": d, **v} for d, v in log.items()],
        key=lambda x: x["date"]
    )
    return jsonify(entries)

@app.route("/api/measurements", methods=["POST"])
def save_measurement():
    d = request.json.get("date", today())
    log = load_json("measurements", {})
    entry = {"logged": datetime.now().isoformat()}
    for f in MEASUREMENT_FIELDS:
        v = request.json.get(f)
        if v is not None and v != "":
            entry[f] = float(v)
    log[d] = entry
    save_json("measurements", log)
    return jsonify({"date": d, **entry})

@app.route("/api/measurements/<string:d>", methods=["DELETE"])
def delete_measurement(d):
    log = load_json("measurements", {})
    log.pop(d, None)
    save_json("measurements", log)
    return jsonify({"ok": True})

@app.route("/api/export")
def export_excel():
    d = request.args.get("date", today())
    tasks     = load_json("tasks", {}).get(d, [])
    habit_list= load_json("habit_list", [])
    habit_log = load_json("habit_log", {}).get(d, {})
    work_list = load_json("work", {}).get(d, [])
    goals     = load_json("goals", [])
    diet_list = load_json("diet", {}).get(d, [])
    notes_list= load_json("notes", {}).get(d, [])
    meas_log  = load_json("measurements", {})
    meas_list = sorted([{"date": k, **v} for k, v in meas_log.items()], key=lambda x: x["date"])
    profile   = load_json("profile", {})
    weight_log= load_json("weight_log", {})

    wb = Workbook()

    # ── color palette ──────────────────────────────────────────────────────
    C = {
        "navy":    "1E3A5F",
        "blue":    "2E86AB",
        "green":   "27AE60",
        "orange":  "F39C12",
        "red":     "E74C3C",
        "purple":  "8E44AD",
        "teal":    "16A085",
        "light":   "EBF5FB",
        "light2":  "F0FFF4",
        "light3":  "FFF9E6",
        "light4":  "F9F0FF",
        "white":   "FFFFFF",
        "gray":    "F8F9FA",
        "dgray":   "6C757D",
    }

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1 – DASHBOARD
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 3

    # title banner
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 10
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value = f"📅  DAILY ACTIVITY TRACKER  —  {d}"
    c.fill = PatternFill("solid", fgColor=C["navy"])
    c.font = Font(color=C["white"], bold=True, size=20)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # summary cards (row 4-8)
    done   = sum(1 for t in tasks if t.get("done"))
    total  = len(tasks)
    task_pct = (done / total * 100) if total else 0

    hab_pct_list = []
    for h in habit_list:
        logged = float(habit_log.get(str(h["id"]), 0))
        target = float(h.get("target", 1))
        hab_pct_list.append(min(logged / target * 100, 100) if target else 0)
    hab_pct = sum(hab_pct_list) / len(hab_pct_list) if hab_pct_list else 0

    total_work = sum(w["hours"] for w in work_list)
    total_cal  = sum(e["calories"] for e in diet_list)

    cards = [
        ("✅  Tasks",    f"{done}/{total} done",   f"{task_pct:.0f}%",  C["blue"],   C["light"]),
        ("💪  Habits",   f"{len(hab_pct_list)} tracked", f"{hab_pct:.0f}% avg", C["green"],  C["light2"]),
        ("⏱  Work",      f"{total_work:.1f} hrs",  "Today",             C["orange"], C["light3"]),
        ("🥗  Calories", f"{total_cal} kcal",       f"{len(diet_list)} meals", C["purple"], C["light4"]),
    ]

    ws.row_dimensions[4].height = 14
    for ci, (title, sub, val, color, bg) in enumerate(cards):
        col = ci + 2  # B=2 C=3 D=4 E=5
        ws.row_dimensions[5].height = 20
        ws.row_dimensions[6].height = 30
        ws.row_dimensions[7].height = 24
        ws.row_dimensions[8].height = 14

        for r in range(5, 9):
            ws.cell(r, col).fill = PatternFill("solid", fgColor=bg)
            ws.cell(r, col).border = thin_border()

        t = ws.cell(5, col, title)
        t.fill = PatternFill("solid", fgColor=color)
        t.font = Font(color=C["white"], bold=True, size=11)
        t.alignment = Alignment(horizontal="center", vertical="center")

        v = ws.cell(6, col, val)
        v.font = Font(color=color, bold=True, size=22)
        v.alignment = Alignment(horizontal="center", vertical="center")

        s = ws.cell(7, col, sub)
        s.font = Font(color=C["dgray"], size=10)
        s.alignment = Alignment(horizontal="center")

    # quick task list
    ws.row_dimensions[10].height = 24
    ws.merge_cells("B10:E10")
    h = ws["B10"]
    h.value = "📝  TODAY'S TASKS OVERVIEW"
    h.fill = PatternFill("solid", fgColor=C["navy"])
    h.font = Font(color=C["white"], bold=True, size=13)
    h.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Task", "Category", "Priority", "Status"]
    for ci, hdr in enumerate(headers):
        c = ws.cell(11, ci + 2, hdr)
        c.fill = PatternFill("solid", fgColor=C["blue"])
        c.font = Font(color=C["white"], bold=True)
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[11].height = 18

    for ri, t in enumerate(tasks):
        row = ri + 12
        ws.row_dimensions[row].height = 18
        status = "✅ Done" if t.get("done") else "⏳ Pending"
        bg = "E8F8F5" if t.get("done") else "FEFEFE"
        vals = [t["text"], t.get("category",""), t.get("priority",""), status]
        for ci, v in enumerate(vals):
            cell = ws.cell(row, ci + 2, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci > 0 else "left",
                                        vertical="center", wrap_text=True)

    add_thick_border(ws, 10, 2, 11 + len(tasks), 5)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2 – TASKS
    # ════════════════════════════════════════════════════════════════════════
    wt = wb.create_sheet("✅ Tasks")
    wt.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFG", [3, 5, 35, 16, 14, 14, 14]):
        wt.column_dimensions[col].width = w

    wt.row_dimensions[1].height = 10
    wt.row_dimensions[2].height = 45
    wt.merge_cells("B2:G2")
    c = wt["B2"]
    c.value = f"✅  TASKS  —  {d}"
    c.fill = PatternFill("solid", fgColor=C["blue"])
    c.font = Font(color=C["white"], bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # stats row
    wt.row_dimensions[4].height = 22
    stats = [("Total", total), ("Completed", done),
             ("Pending", total - done), ("Completion %", f"{task_pct:.1f}%")]
    for i, (lbl, val) in enumerate(stats):
        bg = [C["navy"], C["green"], C["orange"], C["blue"]][i]
        lc = wt.cell(4, i + 2, f"{lbl}: {val}")
        lc.fill = PatternFill("solid", fgColor=bg)
        lc.font = Font(color=C["white"], bold=True, size=11)
        lc.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["#", "Task Description", "Category", "Priority", "Status", "Created"]
    wt.row_dimensions[6].height = 22
    for ci, h in enumerate(headers):
        c = wt.cell(6, ci + 2, h)
        c.fill = PatternFill("solid", fgColor=C["navy"])
        c.font = Font(color=C["white"], bold=True)
        c.alignment = Alignment(horizontal="center")

    pri_colors = {"High": "E74C3C", "Medium": "F39C12", "Low": "27AE60"}
    for ri, t in enumerate(tasks):
        row = ri + 7
        wt.row_dimensions[row].height = 20
        bg = "E8F8F5" if t.get("done") else ("FFF9E6" if t.get("priority") == "High" else "FEFEFE")
        vals = [ri + 1, t["text"], t.get("category","General"),
                t.get("priority","Medium"),
                "✅ Done" if t.get("done") else "⏳ Pending",
                t.get("created","")[:10]]
        for ci, v in enumerate(vals):
            cell = wt.cell(row, ci + 2, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci != 1 else "left",
                                        vertical="center")
            if ci == 3:  # priority
                pc = pri_colors.get(str(v), C["dgray"])
                cell.font = Font(color=pc, bold=True)

    if tasks:
        add_thick_border(wt, 6, 2, 6 + len(tasks), 7)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3 – HABITS
    # ════════════════════════════════════════════════════════════════════════
    wh = wb.create_sheet("💪 Habits")
    wh.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [3, 5, 20, 14, 14, 14, 28, 3]):
        wh.column_dimensions[col].width = w

    wh.row_dimensions[2].height = 45
    wh.merge_cells("B2:G2")
    c = wh["B2"]
    c.value = f"💪  HABITS  —  {d}"
    c.fill = PatternFill("solid", fgColor=C["green"])
    c.font = Font(color=C["white"], bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["", "Habit", "Target", "Logged", "Unit", "Progress"]
    wh.row_dimensions[4].height = 22
    for ci, h in enumerate(headers):
        c = wh.cell(4, ci + 2, h)
        c.fill = PatternFill("solid", fgColor=C["navy"])
        c.font = Font(color=C["white"], bold=True)
        c.alignment = Alignment(horizontal="center")

    for ri, h in enumerate(habit_list):
        row = ri + 5
        wh.row_dimensions[row].height = 24
        logged = float(habit_log.get(str(h["id"]), 0))
        target = float(h.get("target", 1))
        pct = min(logged / target * 100, 100) if target else 0
        bg = "E8F8F5" if pct >= 100 else ("FFF9E6" if pct >= 50 else "FFF0F0")

        vals = [h.get("icon",""), h["name"], target, logged, h.get("unit","")]
        for ci, v in enumerate(vals):
            cell = wh.cell(row, ci + 2, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

        pct_bar(wh, row, 7, pct)
        wh.cell(row, 7).fill = PatternFill("solid", fgColor=bg)
        wh.cell(row, 7).border = thin_border()

    add_thick_border(wh, 4, 2, 4 + len(habit_list), 7)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 4 – WORK HOURS
    # ════════════════════════════════════════════════════════════════════════
    ww = wb.create_sheet("⏱ Work Hours")
    ww.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [3, 30, 14, 36, 3]):
        ww.column_dimensions[col].width = w

    ww.row_dimensions[2].height = 45
    ww.merge_cells("B2:D2")
    c = ww["B2"]
    c.value = f"⏱  WORK HOURS  —  {d}"
    c.fill = PatternFill("solid", fgColor=C["orange"])
    c.font = Font(color=C["white"], bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    total_h = sum(w["hours"] for w in work_list)
    ww.row_dimensions[4].height = 22
    c = ww.cell(4, 2, f"Total Hours Today: {total_h:.1f} hrs")
    c.fill = PatternFill("solid", fgColor=C["navy"])
    c.font = Font(color=C["white"], bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ww.merge_cells("B4:D4")

    headers = ["Project / Task", "Hours", "Notes"]
    ww.row_dimensions[6].height = 22
    for ci, h in enumerate(headers):
        c = ww.cell(6, ci + 2, h)
        c.fill = PatternFill("solid", fgColor=C["navy"])
        c.font = Font(color=C["white"], bold=True)
        c.alignment = Alignment(horizontal="center")

    for ri, w in enumerate(work_list):
        row = ri + 7
        ww.row_dimensions[row].height = 20
        bg = C["light3"] if ri % 2 == 0 else C["white"]
        vals = [w["project"], w["hours"], w.get("notes","")]
        for ci, v in enumerate(vals):
            cell = ww.cell(row, ci + 2, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci == 1 else "left",
                                        vertical="center", wrap_text=True)

    if work_list:
        add_thick_border(ww, 6, 2, 6 + len(work_list), 4)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 5 – GOALS
    # ════════════════════════════════════════════════════════════════════════
    wg = wb.create_sheet("🎯 Goals")
    wg.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGH", [3, 28, 16, 14, 14, 14, 28, 3]):
        wg.column_dimensions[col].width = w

    wg.row_dimensions[2].height = 45
    wg.merge_cells("B2:G2")
    c = wg["B2"]
    c.value = "🎯  GOALS & PROGRESS"
    c.fill = PatternFill("solid", fgColor=C["purple"])
    c.font = Font(color=C["white"], bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Goal", "Category", "Current", "Target", "Deadline", "Progress"]
    wg.row_dimensions[4].height = 22
    for ci, h in enumerate(headers):
        c = wg.cell(4, ci + 2, h)
        c.fill = PatternFill("solid", fgColor=C["navy"])
        c.font = Font(color=C["white"], bold=True)
        c.alignment = Alignment(horizontal="center")

    for ri, g in enumerate(goals):
        row = ri + 5
        wg.row_dimensions[row].height = 24
        cur = float(g.get("current", 0))
        tgt = float(g.get("target", 1))
        pct = min(cur / tgt * 100, 100) if tgt else 0
        bg = "E8F8F5" if pct >= 100 else C["light4"]
        vals = [g["title"], g.get("category",""), cur, tgt, g.get("deadline","")]
        for ci, v in enumerate(vals):
            cell = wg.cell(row, ci + 2, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci > 0 else "left",
                                        vertical="center")
        pct_bar(wg, row, 7, pct)
        wg.cell(row, 7).fill = PatternFill("solid", fgColor=bg)
        wg.cell(row, 7).border = thin_border()

    if goals:
        add_thick_border(wg, 4, 2, 4 + len(goals), 7)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 6 – DIET
    # ════════════════════════════════════════════════════════════════════════
    wd = wb.create_sheet("🥗 Diet")
    wd.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHI", [3, 16, 28, 12, 12, 12, 12, 12, 3]):
        wd.column_dimensions[col].width = w

    wd.row_dimensions[2].height = 45
    wd.merge_cells("B2:H2")
    c = wd["B2"]
    c.value = f"🥗  DIET LOG  —  {d}"
    c.fill = PatternFill("solid", fgColor=C["teal"])
    c.font = Font(color=C["white"], bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # macro totals
    totals = {"calories": 0, "protein": 0, "carbs": 0, "fat": 0}
    for e in diet_list:
        for k in totals:
            totals[k] += e.get(k, 0)

    wd.row_dimensions[4].height = 22
    macro_labels = [
        (f"🔥 {totals['calories']} kcal", C["orange"]),
        (f"💪 {totals['protein']:.1f}g protein", C["blue"]),
        (f"🍞 {totals['carbs']:.1f}g carbs", C["green"]),
        (f"🧈 {totals['fat']:.1f}g fat", C["purple"]),
    ]
    for i, (lbl, color) in enumerate(macro_labels):
        c = wd.cell(4, i + 2, lbl)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(color=C["white"], bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Meal", "Food Item", "Calories", "Protein(g)", "Carbs(g)", "Fat(g)", "Time"]
    wd.row_dimensions[6].height = 22
    for ci, h in enumerate(headers):
        c = wd.cell(6, ci + 2, h)
        c.fill = PatternFill("solid", fgColor=C["navy"])
        c.font = Font(color=C["white"], bold=True)
        c.alignment = Alignment(horizontal="center")

    meal_colors = {
        "Breakfast": "FFF9E6", "Lunch": "E8F8F5",
        "Dinner": "F0F0FF", "Snack": "FFF0F5"
    }
    for ri, e in enumerate(diet_list):
        row = ri + 7
        wd.row_dimensions[row].height = 20
        bg = meal_colors.get(e.get("meal",""), C["white"])
        vals = [e.get("meal",""), e.get("food",""), e.get("calories",0),
                e.get("protein",0), e.get("carbs",0), e.get("fat",0),
                e.get("time","")]
        for ci, v in enumerate(vals):
            cell = wd.cell(row, ci + 2, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center" if ci != 1 else "left",
                                        vertical="center")

    if diet_list:
        add_thick_border(wd, 6, 2, 6 + len(diet_list), 8)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 7 – 7-DAY SUMMARY (work + tasks trend)
    # ════════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("📈 7-Day Summary")
    ws7.sheet_view.showGridLines = False
    ws7.column_dimensions["A"].width = 3
    ws7.column_dimensions["B"].width = 16
    ws7.column_dimensions["C"].width = 16
    ws7.column_dimensions["D"].width = 16
    ws7.column_dimensions["E"].width = 16
    ws7.column_dimensions["F"].width = 3

    ws7.row_dimensions[2].height = 45
    ws7.merge_cells("B2:E2")
    c = ws7["B2"]
    c.value = "📈  7-DAY SUMMARY"
    c.fill = PatternFill("solid", fgColor=C["navy"])
    c.font = Font(color=C["white"], bold=True, size=16)
    c.alignment = Alignment(horizontal="center", vertical="center")

    all_tasks = load_json("tasks", {})
    all_work  = load_json("work", {})

    day_labels, day_done, day_total, day_work = [], [], [], []
    for i in range(6, -1, -1):
        day = (date.today() - timedelta(days=i)).isoformat()
        dt = all_tasks.get(day, [])
        dw = all_work.get(day, [])
        day_labels.append(day[5:])  # MM-DD
        day_done.append(sum(1 for t in dt if t.get("done")))
        day_total.append(len(dt))
        day_work.append(sum(w["hours"] for w in dw))

    headers = ["Date", "Tasks Done", "Total Tasks", "Work Hours"]
    ws7.row_dimensions[4].height = 22
    for ci, h in enumerate(headers):
        c = ws7.cell(4, ci + 2, h)
        c.fill = PatternFill("solid", fgColor=C["navy"])
        c.font = Font(color=C["white"], bold=True)
        c.alignment = Alignment(horizontal="center")

    for ri in range(7):
        row = ri + 5
        ws7.row_dimensions[row].height = 20
        bg = C["light"] if ri % 2 == 0 else C["white"]
        vals = [day_labels[ri], day_done[ri], day_total[ri], day_work[ri]]
        for ci, v in enumerate(vals):
            cell = ws7.cell(row, ci + 2, v)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

    add_thick_border(ws7, 4, 2, 11, 5)

    # bar chart – tasks
    chart = BarChart()
    chart.type = "col"
    chart.title = "Tasks Completed (7 days)"
    chart.y_axis.title = "Count"
    chart.style = 10
    chart.width = 18
    chart.height = 12
    data_ref  = Reference(ws7, min_col=3, max_col=4, min_row=4, max_row=11)
    cats_ref  = Reference(ws7, min_col=2, min_row=5, max_row=11)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws7.add_chart(chart, "B14")

    # line chart – work hours
    lc = LineChart()
    lc.title = "Work Hours (7 days)"
    lc.y_axis.title = "Hours"
    lc.style = 10
    lc.width = 18
    lc.height = 12
    ldata = Reference(ws7, min_col=5, min_row=4, max_row=11)
    lcats = Reference(ws7, min_col=2, min_row=5, max_row=11)
    lc.add_data(ldata, titles_from_data=True)
    lc.set_categories(lcats)
    ws7.add_chart(lc, "B30")

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 8 – JOURNAL / NOTES
    # ════════════════════════════════════════════════════════════════════════
    wn = wb.create_sheet("📓 Journal")
    wn.sheet_view.showGridLines = False
    wn.column_dimensions["A"].width = 12
    wn.column_dimensions["B"].width = 20
    wn.column_dimensions["C"].width = 14
    wn.column_dimensions["D"].width = 14
    wn.column_dimensions["E"].width = 25
    wn.column_dimensions["F"].width = 60

    wn.merge_cells("A1:F1")
    t = wn["A1"]
    t.value = f"📓 Journal & Notes — {d}"
    t.font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C["purple"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    wn.row_dimensions[1].height = 36

    NOTE_COLORS = {
        "journal":   C["blue"],
        "idea":      C["orange"],
        "gratitude": C["purple"],
        "note":      C["green"],
    }
    MOOD_LABELS = {"": "", "😊": "😊", "😐": "😐", "😔": "😔", "🤩": "🤩", "😤": "😤"}

    hdrs = ["#", "Title", "Type", "Mood", "Tags", "Content"]
    hdr_fills = [C["purple"]] * 6
    for ci, (h, fc) in enumerate(zip(hdrs, hdr_fills), 1):
        c = wn.cell(2, ci, h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=fc)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style="medium", color="FFFFFF"))

    if notes_list:
        for ri, note in enumerate(notes_list, 1):
            row = ri + 2
            wn.row_dimensions[row].height = 50
            note_type = note.get("type", "note")
            row_color = NOTE_COLORS.get(note_type, C["green"])

            vals = [
                ri,
                note.get("title") or "(untitled)",
                note_type.capitalize(),
                note.get("mood", ""),
                ", ".join(note.get("tags", [])),
                note.get("content", ""),
            ]
            for ci, val in enumerate(vals, 1):
                c = wn.cell(row, ci, val)
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = Border(
                    left=Side(style="thin", color="DDDDDD"),
                    right=Side(style="thin", color="DDDDDD"),
                    bottom=Side(style="thin", color="DDDDDD"),
                )
                if ci == 1:
                    c.font = Font(name="Calibri", bold=True, color=row_color)
                    c.alignment = Alignment(horizontal="center", vertical="top")
                if ci == 3:
                    c.font = Font(name="Calibri", bold=True, color=row_color)
    else:
        wn.merge_cells("A3:F3")
        emp = wn["A3"]
        emp.value = "No notes recorded for this date."
        emp.font = Font(name="Calibri", italic=True, color="999999")
        emp.alignment = Alignment(horizontal="center", vertical="center")
        wn.row_dimensions[3].height = 30

    # ── Body Measurements sheet ────────────────────────────────────────────
    wm = wb.create_sheet("📏 Measurements")
    wm.sheet_view.showGridLines = False
    wm.column_dimensions["A"].width = 14
    for col, w in zip("BCDEFGHIJ", [10,10,10,10,10,10,10,10,10]):
        wm.column_dimensions[col].width = w
    wm.row_dimensions[1].height = 36
    wm.row_dimensions[2].height = 22

    wm.merge_cells("A1:J1")
    th = wm["A1"]
    th.value = "📏 Body Measurements History"
    th.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    th.fill = PatternFill("solid", fgColor=C["teal"])
    th.alignment = Alignment(horizontal="center", vertical="center")

    meas_headers = ["Date","Weight(kg)","Body Fat%","Waist(cm)","Chest(cm)","Hips(cm)","Neck(cm)","Bicep(cm)","Thigh(cm)","Calf(cm)"]
    meas_keys    = ["date","weight","body_fat","waist","chest","hips","neck","bicep","thigh","calf"]
    for ci, h in enumerate(meas_headers, 1):
        c = wm.cell(2, ci, h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=C["teal"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style="thin", color="FFFFFF"))

    for ri, entry in enumerate(meas_list, 3):
        prev = meas_list[ri-4] if ri > 3 else None
        fill = PatternFill("solid", fgColor=("F2FDFB" if ri%2==1 else "FFFFFF"))
        for ci, key in enumerate(meas_keys, 1):
            val = entry.get(key, "")
            if key == "date":
                cell_val = val
            elif val == "":
                cell_val = "—"
            else:
                raw = float(val) if val != "" else None
                prev_val = float(prev.get(key, 0)) if prev and prev.get(key) else None
                if raw is not None and prev_val:
                    delta = raw - prev_val
                    sign = "+" if delta > 0 else ""
                    cell_val = f"{raw:.1f} ({sign}{delta:.1f})"
                else:
                    cell_val = f"{raw:.1f}" if raw is not None else "—"
            c = wm.cell(ri, ci, cell_val)
            c.font = Font(name="Calibri", size=9, bold=(key=="date"))
            c.fill = fill
            c.alignment = Alignment(horizontal="center" if ci>1 else "left", vertical="center")
        wm.row_dimensions[ri].height = 18

    if not meas_list:
        wm.merge_cells("A3:J3")
        emp = wm["A3"]
        emp.value = "No measurements recorded yet."
        emp.font = Font(name="Calibri", italic=True, color="999999")
        emp.alignment = Alignment(horizontal="center", vertical="center")
        wm.row_dimensions[3].height = 30

    # ── Progress Report sheet ──────────────────────────────────────────────
    wp = wb.create_sheet("🚀 Progress Report")
    wp.sheet_view.showGridLines = False
    wp.column_dimensions["A"].width = 28
    wp.column_dimensions["B"].width = 18
    wp.column_dimensions["C"].width = 18
    wp.column_dimensions["D"].width = 18
    wp.row_dimensions[1].height = 36
    wp.row_dimensions[2].height = 22

    wp.merge_cells("A1:D1")
    ph = wp["A1"]
    ph.value = "🚀 Progress Report"
    ph.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ph.fill = PatternFill("solid", fgColor=C["purple"])
    ph.alignment = Alignment(horizontal="center", vertical="center")

    p_headers = ["Metric", "First Recorded", "Latest", "Change"]
    for ci, h in enumerate(p_headers, 1):
        c = wp.cell(2, ci, h)
        c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=C["purple"])
        c.alignment = Alignment(horizontal="center", vertical="center")

    prog_rows = []
    # Weight from weight log
    wl_sorted = sorted(weight_log.items())
    if wl_sorted:
        first_w = wl_sorted[0][1]["weight"]
        last_w  = wl_sorted[-1][1]["weight"]
        prog_rows.append(("Weight (kg)", f"{first_w:.1f}", f"{last_w:.1f}", f"{last_w-first_w:+.1f}"))
    # Measurements
    if len(meas_list) >= 1:
        first_m, last_m = meas_list[0], meas_list[-1]
        for key, label, unit in [
            ("body_fat","Body Fat","%"),("waist","Waist","cm"),
            ("chest","Chest","cm"),("hips","Hips","cm"),("neck","Neck","cm"),
            ("bicep","Bicep","cm"),("thigh","Thigh","cm"),("calf","Calf","cm"),
        ]:
            fv = first_m.get(key)
            lv = last_m.get(key)
            if fv is not None and lv is not None:
                prog_rows.append((f"{label} ({unit})", f"{float(fv):.1f}", f"{float(lv):.1f}", f"{float(lv)-float(fv):+.1f}"))
    # Goals
    goals_data = load_json("goals", [])
    for g in goals_data:
        pct = min((g.get("current",0)/g["target"])*100,100) if g["target"] else 0
        prog_rows.append((f"Goal: {g['title']}", f"Target: {g['target']} {g.get('unit','')}",
                          f"{g.get('current',0)} {g.get('unit','')}", f"{pct:.1f}%"))

    for ri, row in enumerate(prog_rows, 3):
        fill = PatternFill("solid", fgColor=("F9F0FF" if ri%2==1 else "FFFFFF"))
        for ci, val in enumerate(row, 1):
            c = wp.cell(ri, ci, val)
            c.font = Font(name="Calibri", size=10, bold=(ci==1))
            c.fill = fill
            c.alignment = Alignment(horizontal="center" if ci>1 else "left", vertical="center")
            if ci == 4:
                try:
                    num = float(val.replace("+","").replace("%",""))
                    c.font = Font(name="Calibri", size=10, bold=True,
                                  color=(C["red"] if num > 0 and "fat" in row[0].lower() or
                                         num < 0 and "fat" not in row[0].lower() and "%" not in val
                                         else C["green"]))
                except: pass
        wp.row_dimensions[ri].height = 20

    if not prog_rows:
        wp.merge_cells("A3:D3")
        emp = wp["A3"]
        emp.value = "Log weight and measurements to see your progress here."
        emp.font = Font(name="Calibri", italic=True, color="999999")
        emp.alignment = Alignment(horizontal="center", vertical="center")
        wp.row_dimensions[3].height = 30

    # ── freeze panes & save ────────────────────────────────────────────────
    for sheet in wb.worksheets:
        sheet.freeze_panes = sheet.cell(3, 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"ActivityTracker_{d}.xlsx"
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Wellness / Recovery Tracking ─────────────────────────────────────────────

@app.route("/api/wellness", methods=["GET"])
def get_wellness():
    d = request.args.get("date", today())
    return jsonify(load_json("wellness", {}).get(d, {}))

@app.route("/api/wellness", methods=["POST"])
def save_wellness():
    d = request.json.get("date", today())
    log = load_json("wellness", {})
    entry = {
        "sleep_hours":   float(request.json.get("sleep_hours", 0) or 0),
        "sleep_quality": int(request.json.get("sleep_quality", 5) or 5),
        "energy":        int(request.json.get("energy", 5) or 5),
        "stress":        int(request.json.get("stress", 5) or 5),
        "soreness":      int(request.json.get("soreness", 3) or 3),
        "notes":         request.json.get("notes", ""),
        "logged":        datetime.now().isoformat()
    }
    log[d] = entry
    save_json("wellness", log)
    return jsonify({"date": d, **entry})

@app.route("/api/wellness/history", methods=["GET"])
def get_wellness_history():
    days = int(request.args.get("days", 14))
    log  = load_json("wellness", {})
    end  = date.today()
    out  = []
    for i in range(days-1, -1, -1):
        d = (end - timedelta(days=i)).isoformat()
        if d in log:
            out.append({"date": d, **log[d]})
    return jsonify(out)

# ── Training / Workout Tracking ───────────────────────────────────────────────

@app.route("/api/workouts", methods=["GET"])
def get_workouts():
    d = request.args.get("date", today())
    return jsonify(load_json("workouts", {}).get(d, []))

@app.route("/api/workouts", methods=["POST"])
def add_workout():
    d = request.json.get("date", today())
    workouts = load_json("workouts", {})
    entry = {
        "id":           int(datetime.now().timestamp() * 1000),
        "name":         request.json.get("name", "Workout"),
        "type":         request.json.get("type", "strength"),
        "duration_mins":int(request.json.get("duration_mins", 0) or 0),
        "exercises":    request.json.get("exercises", []),
        "notes":        request.json.get("notes", ""),
        "logged":       datetime.now().isoformat()
    }
    workouts.setdefault(d, []).append(entry)
    save_json("workouts", workouts)
    return jsonify({"date": d, **entry})

@app.route("/api/workouts/<int:wid>", methods=["DELETE"])
def delete_workout(wid):
    d = request.args.get("date", today())
    workouts = load_json("workouts", {})
    workouts[d] = [w for w in workouts.get(d, []) if w["id"] != wid]
    save_json("workouts", workouts)
    return jsonify({"ok": True})

@app.route("/api/workouts/week", methods=["GET"])
def get_workouts_week():
    out = []
    for i in range(6, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        sessions = load_json("workouts", {}).get(d, [])
        out.append({"date": d, "sessions": sessions, "count": len(sessions),
                    "total_mins": sum(s.get("duration_mins", 0) for s in sessions)})
    return jsonify(out)

@app.route("/api/workouts/all", methods=["GET"])
def get_all_workouts():
    workouts = load_json("workouts", {})
    result = []
    for d in sorted(workouts.keys(), reverse=True):
        for w in workouts[d]:
            result.append({**w, "date": d})
    return jsonify(result)

# ── Elite AI Analysis Engine ──────────────────────────────────────────────────

def _weight_trend_analysis(weight_log):
    sorted_entries = sorted(weight_log.items())
    if len(sorted_entries) < 2:
        return {"status": "insufficient_data", "rate": 0, "trend": "unknown", "current": None}
    cutoff = (date.today() - timedelta(days=14)).isoformat()
    recent = [(d, v["weight"]) for d, v in sorted_entries if d >= cutoff]
    if len(recent) < 2:
        recent = sorted_entries[-min(7, len(sorted_entries)):]
    n = len(recent)
    xs = list(range(n)); ys = [r[1] for r in recent]
    denom = n*sum(x*x for x in xs) - sum(xs)**2
    slope = (n*sum(x*y for x,y in zip(xs,ys)) - sum(xs)*sum(ys)) / denom if denom else 0
    weekly = slope * 7
    return {
        "rate":         round(weekly, 3),
        "current":      recent[-1][1],
        "start":        recent[0][1],
        "total_change": round(recent[-1][1] - recent[0][1], 2),
        "trend":        "losing" if weekly < -0.15 else "gaining" if weekly > 0.15 else "stable",
        "entries":      n
    }

def _diet_adherence_analysis(diet, diet_plan, days=7):
    plan_cal = diet_plan.get("calories", 2000)
    plan_pro = diet_plan.get("protein", 150)
    tracked, total_cal_pct, total_pro = 0, 0, 0
    for i in range(days):
        d = (date.today() - timedelta(days=i)).isoformat()
        entries = diet.get(d, [])
        if entries:
            tracked += 1
            day_cal = sum(e.get("calories", 0) for e in entries)
            day_pro = sum(e.get("protein",  0) for e in entries)
            total_cal_pct += day_cal / plan_cal if plan_cal else 0
            total_pro += day_pro
    avg_cal_pct = round(total_cal_pct / tracked * 100) if tracked else 0
    avg_pro     = round(total_pro / tracked) if tracked else 0
    return {"tracked": tracked, "days": days,
            "adherence_pct": round(tracked/days*100),
            "avg_cal_pct": avg_cal_pct, "avg_protein": avg_pro,
            "plan_protein": plan_pro}

def _recovery_analysis(wellness_log, days=7):
    scores, sleep_hrs, energies, stresses = [], [], [], []
    for i in range(days):
        d = (date.today() - timedelta(days=i)).isoformat()
        w = wellness_log.get(d, {})
        if w:
            sh = float(w.get("sleep_hours", 0) or 0)
            sq = int(w.get("sleep_quality", 5) or 5)
            en = int(w.get("energy",        5) or 5)
            st = int(w.get("stress",        5) or 5)
            so = int(w.get("soreness",      3) or 3)
            sleep_score = min(sh / 8 * 10, 10)
            day_score   = (sleep_score + sq + en + (10-st) + (10-so)) / 5
            scores.append(day_score); sleep_hrs.append(sh)
            energies.append(en); stresses.append(st)
    if not scores:
        return {"score": None, "status": "no_data", "tracked": 0}
    avg = sum(scores)/len(scores)
    status = "excellent" if avg>=8 else "good" if avg>=6.5 else "moderate" if avg>=5 else "poor"
    return {"score": round(avg*10), "status": status, "tracked": len(scores),
            "avg_sleep": round(sum(sleep_hrs)/len(sleep_hrs),1),
            "avg_energy": round(sum(energies)/len(energies),1),
            "avg_stress": round(sum(stresses)/len(stresses),1)}

def _training_analysis(workouts_log, days=7):
    sessions, total_mins, types = 0, 0, {}
    for i in range(days):
        d = (date.today() - timedelta(days=i)).isoformat()
        for w in workouts_log.get(d, []):
            sessions += 1
            total_mins += w.get("duration_mins", 0)
            t = w.get("type","strength")
            types[t] = types.get(t,0) + 1
    return {"sessions": sessions, "total_mins": total_mins,
            "avg_session": round(total_mins/sessions) if sessions else 0,
            "dominant_type": max(types, key=types.get) if types else None}

def _bfpct_trend(measurements):
    sorted_m = sorted(measurements.items())
    entries = [(d, v.get("body_fat")) for d,v in sorted_m if v.get("body_fat") is not None]
    if len(entries) < 2:
        return None
    return round(entries[-1][1] - entries[0][1], 1)

@app.route("/api/ai/analysis")
def ai_analysis():
    profile       = load_json("profile", {})
    weight_log    = load_json("weight_log", {})
    measurements  = load_json("measurements", {})
    diet          = load_json("diet", {})
    diet_plan     = load_json("diet_plan", {"calories":2000,"protein":150,"carbs":250,"fat":65})
    wellness_log  = load_json("wellness", {})
    workouts_log  = load_json("workouts", {})
    goals_list    = load_json("goals", [])

    goal     = profile.get("goal", "maintain")
    wt       = _weight_trend_analysis(weight_log)
    adh      = _diet_adherence_analysis(diet, diet_plan)
    rec      = _recovery_analysis(wellness_log)
    training = _training_analysis(workouts_log)
    bf_delta = _bfpct_trend(measurements)
    plan_cal = diet_plan.get("calories", 2000)
    plan_pro = diet_plan.get("protein", 150)
    bw       = profile.get("weight", 75)

    insights, alerts, adjustments, today_focus = [], [], {}, []

    # ── Plateau detection ─────────────────────────────────────────
    if wt["entries"] >= 4:
        if goal == "lose_weight" and wt["trend"] == "stable":
            insights.append({"type":"plateau","priority":1,"icon":"🔄",
                "title":"Weight Loss Plateau",
                "body":f"Weight has been stable ({wt['rate']:+.2f} kg/wk) for the past 2 weeks. Your body has adapted — time to break through.",
                "actions":["Cut 150 kcal/day (remove one snack or reduce portion)", "Add 20 min walking after dinner 4×/week", "Audit hidden calories: sauces, drinks, cooking oils", "Consider a 24-hr refeed at maintenance calories to reset leptin"],
                "badge":"warning"})
            adjustments["calories"] = -150
            alerts.append({"type":"warning","msg":"⚠️ Weight plateau — calorie adjustment recommended"})

        elif goal == "gain_muscle" and wt["trend"] == "stable":
            insights.append({"type":"plateau","priority":1,"icon":"💪",
                "title":"Muscle Gain Stalled",
                "body":"Scale hasn't moved in 2 weeks. You need a larger caloric surplus to drive muscle protein synthesis.",
                "actions":["Add 200 kcal from carbs around training (rice, oats, banana)", "Increase training volume by 1 set per muscle group", "Confirm progressive overload — increase weight or reps weekly"],
                "badge":"warning"})
            adjustments["calories"] = 200; adjustments["protein"] = 10
            alerts.append({"type":"info","msg":"ℹ️ Bulk plateau — increase surplus by 200 kcal"})

        elif goal == "lose_weight" and wt["rate"] < -1.2:
            insights.append({"type":"warning","priority":1,"icon":"⚡",
                "title":"Losing Weight Too Fast",
                "body":f"Current rate: {abs(wt['rate']):.1f} kg/week. Safe fat loss is 0.5–1.0 kg/week. Faster than this risks muscle loss and metabolic slowdown.",
                "actions":["Increase calories by 200–250 kcal/day", "Hit minimum 2.0 g protein/kg bodyweight", "Add one resistance session per week to preserve muscle"],
                "badge":"danger"})
            adjustments["calories"] = 250
            alerts.append({"type":"danger","msg":"🚨 Losing weight too fast — increase calories"})

    # ── Poor adherence ────────────────────────────────────────────
    if adh["adherence_pct"] < 50:
        insights.append({"type":"adherence","priority":2,"icon":"📋",
            "title":"Low Tracking Consistency",
            "body":f"Only {adh['tracked']}/{adh['days']} days logged. Consistent tracking is the single highest-leverage habit for transformation.",
            "actions":["Log meals immediately — not at end of day", "Meal prep Sunday to remove decision fatigue", "Set a 8pm phone reminder: 'Did you log today?'", "Start with just logging dinner to build the habit"],
            "badge":"warning"})
    elif adh["adherence_pct"] >= 85:
        insights.append({"type":"success","priority":5,"icon":"🌟",
            "title":"Excellent Tracking Consistency",
            "body":f"You've logged {adh['tracked']}/{adh['days']} days — elite-level consistency. This is what separates people who get results from those who don't.",
            "actions":["Keep this streak going","Review weekly trends to spot optimisation opportunities"],
            "badge":"success"})

    # ── Calorie target adherence ───────────────────────────────────
    if adh["tracked"] >= 3:
        if adh["avg_cal_pct"] < 80:
            insights.append({"type":"nutrition","priority":2,"icon":"🍽️",
                "title":"Chronic Undereating Detected",
                "body":f"Averaging {adh['avg_cal_pct']}% of your {plan_cal} kcal target. Sustained undereating downregulates metabolism, increases cortisol, and breaks down muscle tissue.",
                "actions":["Add a 300–400 kcal nutrient-dense snack (nuts, Greek yogurt, banana + PB)", "Front-load calories — eat more at breakfast and lunch", "Prepare meals in advance so food is ready when hunger hits"],
                "badge":"warning"})
        elif adh["avg_cal_pct"] > 120:
            insights.append({"type":"nutrition","priority":2,"icon":"🍽️",
                "title":"Consistently Overeating",
                "body":f"Averaging {adh['avg_cal_pct']}% of your calorie target. Sustained surplus beyond goal will increase fat storage.",
                "actions":["Use a food scale for calorie-dense foods (oils, nuts, cheese)", "Eat slowly — it takes 20 min for fullness signals to reach your brain", "Replace one high-calorie snack with a high-volume, low-calorie option"],
                "badge":"warning"})

    # ── Protein adequacy ─────────────────────────────────────────
    if adh["tracked"] >= 3 and adh["avg_protein"] < plan_pro * 0.75:
        insights.append({"type":"protein","priority":2,"icon":"💪",
            "title":"Protein Target Missed",
            "body":f"Averaging {adh['avg_protein']}g vs your {plan_pro}g target. Protein is the most critical macro — it drives muscle synthesis and keeps you full.",
            "actions":[f"Add a protein source to every meal (target 30–40g per meal)", "Greek yogurt + whey shake = easy 50g boost", "Replace carb-heavy snacks with high-protein alternatives"],
            "badge":"warning"})

    # ── Recovery ─────────────────────────────────────────────────
    if rec["score"] is not None:
        if rec["score"] < 45:
            insights.append({"type":"recovery","priority":1,"icon":"😴",
                "title":"Recovery in Critical Zone",
                "body":f"Recovery score: {rec['score']}/100. Training hard while under-recovered accelerates overtraining, injury risk, and hormonal disruption.",
                "actions":["Take a full rest day or active recovery (walk/stretch only)", f"Target {max(7.5, rec.get('avg_sleep',6)+1):.0f}+ hours of sleep tonight", "Reduce training intensity by 30% this week", "Prioritise magnesium-rich foods and consider 400mg magnesium glycinate before bed"],
                "badge":"danger"})
            alerts.append({"type":"danger","msg":"🚨 Recovery critical — rest day strongly recommended"})
        elif rec["score"] < 65:
            insights.append({"type":"recovery","priority":2,"icon":"🔋",
                "title":"Suboptimal Recovery",
                "body":f"Recovery score: {rec['score']}/100. You're accumulating fatigue. Address sleep and stress before adding more training load.",
                "actions":["Aim for 7.5–8.5 hours sleep", "Reduce high-stress activities", "Add 10 min stretching or foam rolling post-workout"],
                "badge":"info"})
        elif rec["score"] >= 80:
            insights.append({"type":"success","priority":5,"icon":"⚡",
                "title":"Recovery Excellent",
                "body":f"Recovery score: {rec['score']}/100. You're well-rested and ready to train hard.",
                "actions":["Capitalise with a high-intensity session today", "This is an ideal day to attempt personal records"],
                "badge":"success"})

    # ── Training frequency ────────────────────────────────────────
    if training["sessions"] == 0:
        insights.append({"type":"training","priority":3,"icon":"🏋️",
            "title":"No Training Logged This Week",
            "body":"Exercise is a non-negotiable pillar for body composition and metabolic health. Even 3 sessions per week produces significant results.",
            "actions":["Schedule 3 training sessions right now — put them in your calendar", "Start with 30–45 min full-body sessions if time is limited", "Even 20 min of resistance training beats no training"],
            "badge":"warning"})
    elif training["sessions"] >= 5 and rec.get("score", 100) < 60:
        insights.append({"type":"overtraining","priority":2,"icon":"⚠️",
            "title":"High Training Volume + Poor Recovery",
            "body":f"{training['sessions']} sessions this week but recovery is low. More is not always better — super-compensation requires adequate rest.",
            "actions":["Replace one session with active recovery this week", "Ensure at least 48h between training the same muscle groups", "Increase sleep by 30–60 min to support adaptation"],
            "badge":"warning"})

    # ── Body composition (if measurements available) ──────────────
    if bf_delta is not None:
        if goal == "lose_weight" and bf_delta < -1.5:
            insights.append({"type":"body_comp","priority":4,"icon":"📉",
                "title":"Body Fat Trending Down",
                "body":f"Body fat reduced by {abs(bf_delta):.1f}% — excellent progress. This suggests you're losing fat while managing muscle retention.",
                "actions":["Maintain current protein intake to protect lean mass", "Continue current training stimulus"],
                "badge":"success"})
        elif goal == "gain_muscle" and bf_delta > 3:
            insights.append({"type":"body_comp","priority":3,"icon":"📊",
                "title":"Fat Gain Alongside Muscle",
                "body":f"Body fat up {bf_delta:.1f}% during bulk. Some fat gain is normal, but consider a mini cut if this continues.",
                "actions":["Tighten caloric surplus to +200 kcal (cleaner bulk)", "Prioritise compound lifts to maximise muscle stimulus per calorie", "Consider a 2-week mini cut if fat gain continues"],
                "badge":"info"})

    # ── Goal progress ─────────────────────────────────────────────
    completed_goals = [g for g in goals_list if g.get("target",1)>0 and g.get("current",0)/g["target"]>=1]
    if completed_goals:
        for g in completed_goals[:2]:
            insights.append({"type":"achievement","priority":4,"icon":"🏆",
                "title":f"Goal Achieved: {g['title']}",
                "body":f"You've hit {g['current']} {g.get('unit','')} — your target of {g['target']} {g.get('unit','')}. Outstanding.",
                "actions":["Set your next progressive goal", "Celebrate this milestone — it matters"],
                "badge":"success"})

    # ── Overall score ─────────────────────────────────────────────
    score = 50
    score += (adh["adherence_pct"] / 100 - 0.5) * 30
    if rec["score"] is not None:
        score += (rec["score"] / 100 - 0.5) * 25
    if wt["trend"] != "unknown":
        on_track = (goal=="lose_weight" and wt["trend"]=="losing") or \
                   (goal=="gain_muscle" and wt["trend"]=="gaining") or \
                   (goal=="maintain"    and wt["trend"]=="stable")
        score += 15 if on_track else -5
    if training["sessions"] >= 3:
        score += 10
    overall = max(5, min(99, round(score)))

    # ── Grade ─────────────────────────────────────────────────────
    grade = "S" if overall>=90 else "A" if overall>=80 else "B" if overall>=70 else "C" if overall>=55 else "D"

    # ── Today's Focus ─────────────────────────────────────────────
    # Pull top-priority action from top-priority insights
    critical = [i for i in insights if i.get("badge") in ("danger","warning") and i["priority"]<=2]
    for ins in sorted(critical, key=lambda x: x["priority"])[:3]:
        if ins.get("actions"):
            today_focus.append({"icon":ins["icon"],"text":ins["actions"][0],"priority":ins["badge"]})
    # Fill with goal-specific defaults if under 3
    defaults = {
        "lose_weight": [{"icon":"🔥","text":f"Hit your {plan_cal} kcal target — not under","priority":"info"},
                        {"icon":"💧","text":"Drink 2.5–3L of water","priority":"info"},
                        {"icon":"🚶","text":"Hit 8,000+ steps today","priority":"info"}],
        "gain_muscle": [{"icon":"💪","text":f"Eat {plan_pro}g protein — prioritise this above everything","priority":"info"},
                        {"icon":"🏋️","text":"Complete your training session with progressive overload","priority":"info"},
                        {"icon":"😴","text":"Sleep 8+ hours — growth happens at rest","priority":"info"}],
        "maintain":    [{"icon":"⚖️","text":"Log all meals to stay accountable","priority":"info"},
                        {"icon":"🏃","text":"Complete today's planned activity","priority":"info"},
                        {"icon":"💧","text":"Stay hydrated — minimum 2L water","priority":"info"}],
        "athletic":    [{"icon":"⚡","text":"Fuel workouts with adequate carbohydrates","priority":"info"},
                        {"icon":"🏋️","text":"Execute training with full intensity","priority":"info"},
                        {"icon":"🔋","text":"Prioritise recovery — it's where adaptation happens","priority":"info"}],
    }
    for d in defaults.get(goal, defaults["maintain"]):
        if len(today_focus) < 3:
            today_focus.append(d)

    return jsonify({
        "overall_score":  overall,
        "grade":          grade,
        "insights":       sorted(insights, key=lambda x: x.get("priority",5)),
        "alerts":         alerts,
        "adjustments":    adjustments,
        "today_focus":    today_focus[:3],
        "weight_trend":   wt,
        "adherence":      adh,
        "recovery":       rec,
        "training":       training,
        "recommended_calories": plan_cal + adjustments.get("calories", 0),
        "recommended_protein":  plan_pro + adjustments.get("protein",  0),
    })

if __name__ == "__main__":
    app.run(debug=True, port=5050)
